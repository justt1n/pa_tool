import os
from typing import List, Dict

import openpyxl
from pydantic import BaseModel

from constants import TEMPLATE_FOLDER


class CurrencyTemplate(BaseModel):
    action: str = "Sell"
    game: str
    server: str
    faction: str
    currency_per_unit: float
    total_units: float
    minimum_unit_per_order: float
    price_currency: str = "USD"
    price_per_unit: float
    ValueForDiscount: str
    discount: str
    title: str
    duration: int
    delivery_guarantee: int
    delivery_method: str = "Face to Face"
    delivery_character: str = ""
    delivery_instructions: str = ""
    description: str

    class Config:
        arbitrary_types_allowed = True


class ItemTemplate(BaseModel):
    game: str
    server: str
    faction: str
    item_category1: str
    item_category2: str
    item_category3: str
    item_per_unit: float
    unit_price: float
    min_unit_per_order: float
    price_currency: str = "USD"
    ValueForDiscount: float
    discount: float
    offer_duration: int
    delivery_guarantee: int
    delivery_info: str
    cover_image: str
    title: str
    description: str

    class Config:
        arbitrary_types_allowed = True


def currency_templates_to_dicts(templates: List[CurrencyTemplate]) -> List[Dict[str, any]]:
    return [template.dict() for template in templates]


def read_xlsx_file(file_path: str) -> List[Dict[str, any]]:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []

    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_data)

    return data


def write_xlsx_file(file_path: str, data: List[Dict[str, any]]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if not data:
        workbook.save(file_path)
        return

    headers = data[0].keys()
    sheet.append(list(headers))

    for row_data in data:
        sheet.append(list(row_data.values()))

    workbook.save(file_path)
    return True


def load_template(template_name):
    template_path = os.path.join(TEMPLATE_FOLDER, template_name)
    with open(template_path, 'r') as file:
        template_content = file.read()
    return template_content
